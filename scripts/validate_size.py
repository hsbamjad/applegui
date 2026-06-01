"""
scripts/validate_size.py
=========================
Validation script for the apple size estimation pipeline.

Runs the calibrated SizeCalibrator on a recorded video session, matches
committed apple detections to ground-truth caliper measurements, and reports:

  - R2 (predicted vs actual diameter)
  - RMSE (mm)
  - MAE  (mm)
  - Bias (mm, positive = over-estimate)
  - P vs D scatter plot (predicted vs GT diameter)

Target accuracy from Mizushima & Lu (2013):  R2 > 0.99, RMSE < 1.79 mm

Usage
-----
    python scripts/validate_size.py \\
        --videos videos/Source0/G2/G2.avi videos/Source1/G2/G2.avi videos/Source2/G2/G2.avi \\
        --gt data/gt.xlsx \\
        --group G2 \\
        --model models/best.pt \\
        --config config/config.yaml \\
        --output docs/validation_G2.png

Notes
-----
  - Use a DIFFERENT group from the calibration group for true validation.
  - If size_estimation.calibration.small_coeffs is null in config the
    script will exit with a clear message.
"""

from __future__ import annotations

import argparse
import logging
import sys
import warnings
from pathlib import Path
from typing import List

import cv2
import numpy as np

logging.getLogger("ultralytics").setLevel(logging.WARNING)
warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from gui.workers.tracker import AppleTracker, GradeRecord
from gui.workers.size_calibration import SizeCalibrator

log = logging.getLogger(__name__)


# =============================================================================
# Helpers (shared with calibrate_size.py)
# =============================================================================

def load_gt(xlsx_path: str, groups: List[str]) -> List[float]:
    """Load GT average diameters for the specified group(s)."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed. Run: pip install openpyxl")

    wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws   = wb.active
    diameters: List[float] = []
    current_group = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        grp, apple_id, d1, d2 = row[0], row[1], row[2], row[3]
        if grp:
            current_group = str(grp).strip()
        if current_group in groups and apple_id is not None and d1 is not None and d2 is not None:
            diameters.append((float(d1) + float(d2)) / 2.0)

    return diameters


def _prepare_input(ch1: np.ndarray, ch2, ch3, mode: str = "rb-nir1") -> np.ndarray:
    def _gray(f):
        if f is None:
            return np.zeros(ch1.shape[:2], dtype=np.uint8)
        if f.dtype != np.uint8:
            f = cv2.normalize(f, None, 0, 255, cv2.NORM_MINMAX, cv2.CV_8U)
        return f[:, :, 0] if f.ndim == 3 else f

    if mode == "rb-nir1":
        R, B = ch1[:, :, 2], ch1[:, :, 0]
        return np.stack([R, B, _gray(ch2)], axis=2)
    elif mode == "rg-nir1":
        R, G = ch1[:, :, 2], ch1[:, :, 1]
        return np.stack([R, G, _gray(ch2)], axis=2)
    else:
        return ch1


def run_tracker(
    paths: List[str],
    model_path: str,
    calibrator: SizeCalibrator,
    orientation: str,
    n_lanes: int,
    input_mode: str,
    conf_thr: float,
    iou_thr: float,
    device: str,
    min_dpx: float,
    skip_n: int,
) -> List[GradeRecord]:
    """Run YOLO+AppleTracker and return committed GradeRecords."""
    from ultralytics import YOLO

    print(f"\nLoading model: {model_path}")
    model = YOLO(model_path)

    caps = [cv2.VideoCapture(p) for p in paths]
    for cap, p in zip(caps, paths):
        if not cap.isOpened():
            sys.exit(f"Cannot open video: {p}")
    print(f"Videos: {[Path(p).name for p in paths]}")

    h = int(caps[0].get(cv2.CAP_PROP_FRAME_HEIGHT))
    w = int(caps[0].get(cv2.CAP_PROP_FRAME_WIDTH))
    total = int(caps[0].get(cv2.CAP_PROP_FRAME_COUNT))

    tracker = AppleTracker(
        n_lanes         = n_lanes,
        orientation     = orientation,
        exit_frac       = 0.85,
        band_half_frac  = 0.025,
        entry_frac      = 0.35,
        min_frames      = 5,
        size_calibrator = calibrator,
    )

    grades: List[GradeRecord] = []
    frame_shape = (h, w, 3)
    frame_no = 0

    while True:
        frames = [cap.read() for cap in caps]
        oks    = [ok for ok, _ in frames]
        frs    = [fr for _, fr in frames]

        if not oks[0]:
            break

        frame_no += 1
        if frame_no % max(1, skip_n) != 0:
            continue

        if (frame_no // skip_n) % 50 == 0:
            pct = 100 * frame_no / max(total, 1)
            print(f"  Frame {frame_no}/{total}  ({pct:.0f}%)  grades: {len(grades)}")

        ch1 = frs[0] if frs[0] is not None else np.zeros((h, w, 3), np.uint8)
        ch2 = frs[1] if len(frs) > 1 else None
        ch3 = frs[2] if len(frs) > 2 else None
        composite = _prepare_input(ch1, ch2, ch3, input_mode)

        results = model.track(
            source  = composite,
            tracker = "bytetrack.yaml",
            persist = True,
            conf    = conf_thr,
            iou     = iou_thr,
            device  = device,
            imgsz   = 640,
            verbose = False,
            save    = False,
        )

        _, graded = tracker.update(results[0], frame_shape)
        for rec in graded:
            if rec.size_px is not None and rec.size_px >= min_dpx:
                grades.append(rec)

    for cap in caps:
        cap.release()

    print(f"\nDone. {frame_no} frames, {len(grades)} apples committed.")
    return grades


# =============================================================================
# Validation metrics + plot
# =============================================================================

def validate(predicted: List[float], actual: List[float]) -> dict:
    p = np.array(predicted)
    a = np.array(actual)
    err = p - a
    ss_res = np.sum(err ** 2)
    ss_tot = np.sum((a - a.mean()) ** 2)
    r2     = 1 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return {
        "n":    len(p),
        "r2":   float(r2),
        "rmse": float(np.sqrt(np.mean(err ** 2))),
        "mae":  float(np.mean(np.abs(err))),
        "bias": float(np.mean(err)),
    }


def make_pv_d_plot(predicted, actual, metrics, output_path: str) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("matplotlib not installed - skipping plot")
        return

    p = np.array(predicted)
    a = np.array(actual)

    fig, ax = plt.subplots(figsize=(8, 7))
    ax.scatter(a, p, alpha=0.75, s=70, color="#3B82F6",
               edgecolors="white", linewidth=0.5, zorder=3, label="Measured apples")

    # 1:1 line
    lo, hi = min(min(a), min(p)) - 2, max(max(a), max(p)) + 2
    ax.plot([lo, hi], [lo, hi], "r--", linewidth=1.5, label="1:1 line")

    # ±2 mm band
    ax.fill_between([lo, hi], [lo-2, hi-2], [lo+2, hi+2],
                    alpha=0.12, color="red", label="±2 mm band")

    ax.set_xlim(lo, hi)
    ax.set_ylim(lo, hi)
    ax.set_xlabel("GT Diameter (mm)", fontsize=12)
    ax.set_ylabel("Predicted Diameter (mm)", fontsize=12)
    ax.set_title("Apple Size Validation: Predicted vs Ground Truth", fontsize=13, fontweight="bold")
    ax.legend(fontsize=10)
    ax.grid(alpha=0.3)

    # Metrics box
    txt = (f"n = {metrics['n']}\n"
           f"R² = {metrics['r2']:.4f}\n"
           f"RMSE = {metrics['rmse']:.2f} mm\n"
           f"MAE  = {metrics['mae']:.2f} mm\n"
           f"Bias = {metrics['bias']:+.2f} mm")
    ax.text(0.04, 0.96, txt, transform=ax.transAxes, fontsize=10,
            verticalalignment="top",
            bbox=dict(boxstyle="round,pad=0.5", facecolor="white", alpha=0.85))

    # Target thresholds annotation
    target_r2   = 0.99
    target_rmse = 1.79
    r2_ok   = metrics["r2"]   >= target_r2
    rmse_ok = metrics["rmse"] <= target_rmse
    status = "PASS" if (r2_ok and rmse_ok) else "FAIL"
    colour = "#10B981" if status == "PASS" else "#EF4444"
    ax.text(0.96, 0.04, f"Target: R²>{target_r2}  RMSE<{target_rmse}mm\n{status}",
            transform=ax.transAxes, fontsize=10, ha="right", va="bottom",
            color=colour, fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white", alpha=0.85))

    plt.tight_layout()
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    print(f"Plot saved to: {output_path}")


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Validate calibrated size estimation against GT caliper measurements."
    )
    ap.add_argument("--videos",  nargs="+", required=True)
    ap.add_argument("--gt",      required=True)
    ap.add_argument("--groups",  nargs="+", default=["G2"],
                    help="GT session group(s) to validate against (use different group from calibration)")
    ap.add_argument("--model",   required=True)
    ap.add_argument("--config",  default="config/config.yaml")
    ap.add_argument("--output",  default="docs/validation.png")
    ap.add_argument("--orientation", default="LR")
    ap.add_argument("--lanes",   type=int, default=3)
    ap.add_argument("--input-mode", default="rb-nir1")
    ap.add_argument("--conf",    type=float, default=0.5)
    ap.add_argument("--iou",     type=float, default=0.45)
    ap.add_argument("--device",  default="cuda")
    ap.add_argument("--min-dpx", type=float, default=80)
    ap.add_argument("--skip-n",  type=int, default=1)
    args = ap.parse_args()

    print("=" * 60)
    print("APPLE SIZE VALIDATION")
    print("=" * 60)

    # 1. Load calibrated SizeCalibrator from config
    import yaml
    with open(args.config) as f:
        cfg = yaml.safe_load(f)

    size_cfg = cfg.get("size_estimation", {})
    calibrator = SizeCalibrator.from_config(size_cfg)
    if not calibrator.is_ready():
        sys.exit(
            "ERROR: Calibration coefficients are null in config.yaml.\n"
            "Run calibrate_size.py first, then retry."
        )
    print(f"Calibrator loaded (ready={calibrator.is_ready()})")

    # 2. Load GT
    print("\nLoading GT data...")
    gt_diameters = load_gt(args.gt, args.groups)
    if not gt_diameters:
        sys.exit(f"No GT data for groups {args.groups}")
    print(f"  {len(gt_diameters)} apples  "
          f"range: {min(gt_diameters):.1f}-{max(gt_diameters):.1f} mm  "
          f"mean: {np.mean(gt_diameters):.1f} mm")

    # 3. Run tracker with calibrator
    grades = run_tracker(
        paths      = args.videos,
        model_path = args.model,
        calibrator = calibrator,
        orientation = args.orientation,
        n_lanes    = args.lanes,
        input_mode = args.input_mode,
        conf_thr   = args.conf,
        iou_thr    = args.iou,
        device     = args.device,
        min_dpx    = args.min_dpx,
        skip_n     = args.skip_n,
    )

    if not grades:
        sys.exit("No apples committed. Check video, model, and tracker settings.")

    # 4. Match by order
    n_match = min(len(grades), len(gt_diameters))
    if len(grades) != len(gt_diameters):
        print(f"\nWARNING: {len(grades)} graded vs {len(gt_diameters)} GT. Using {n_match}.")

    predicted = []
    actual    = []

    print(f"\n{'#':>4}  {'cx':>7}  {'D_px':>7}  {'D_pred':>8}  {'D_GT':>7}  {'err':>7}")
    print("-" * 50)
    for i in range(n_match):
        rec = grades[i]
        dgt = gt_diameters[i]
        if rec.size_mm is None:
            print(f"{i+1:>4}  size_mm=None (skipped)")
            continue
        err = rec.size_mm - dgt
        print(f"{i+1:>4}  {rec.cx_peak or 0:>7.0f}  {rec.size_px or 0:>7.1f}  "
              f"{rec.size_mm:>8.1f}  {dgt:>7.1f}  {err:>+7.2f}")
        predicted.append(rec.size_mm)
        actual.append(dgt)

    if len(predicted) < 2:
        sys.exit("Not enough matched data points to compute metrics.")

    # 5. Metrics
    m = validate(predicted, actual)

    print("\n" + "=" * 60)
    print("VALIDATION RESULTS")
    print("=" * 60)
    print(f"  n    = {m['n']}")
    print(f"  R2   = {m['r2']:.4f}  (target: > 0.99)")
    print(f"  RMSE = {m['rmse']:.2f} mm  (target: < 1.79 mm)")
    print(f"  MAE  = {m['mae']:.2f} mm")
    print(f"  Bias = {m['bias']:+.2f} mm")

    pass_r2   = m["r2"]   >= 0.99
    pass_rmse = m["rmse"] <= 1.79
    overall   = "PASS" if (pass_r2 and pass_rmse) else "FAIL"
    print(f"\n  Overall: {overall}")
    if not pass_r2:
        print(f"    R2 {m['r2']:.4f} < 0.99  (FAIL)")
    if not pass_rmse:
        print(f"    RMSE {m['rmse']:.2f} mm > 1.79 mm  (FAIL)")

    # 6. Plot
    make_pv_d_plot(predicted, actual, m, args.output)

    print("\nValidation complete.")


if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING, format="%(message)s")
    main()
