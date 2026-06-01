"""
scripts/calibrate_size.py
==========================
Primary calibration script for the apple size estimation pipeline.

Fits the empirical position-aware scale function from Mizushima & Lu (2013):
    N_px(x) = a*x^2 + b*x + c      [pixel diameter of a d_ref-mm object at column x]
    r(x)    = d_ref / N_px(x)       [mm per pixel at column x]

Uses real apple ground-truth (GT) caliper measurements matched to YOLO detections
from recorded video sessions. No calibration balls required.

Usage (one group at a time)
----------------------------
    python scripts/calibrate_size.py \\
        --videos videos/Source0/G1/G1.mp4 videos/Source1/G1/G1.avi videos/Source2/G1/G1.avi \\
        --gt data/gt.xlsx \\
        --group G1 \\
        --model models/best.pt \\
        --config config/config.yaml \\
        --output docs/calibration_G1.png

Output
------
  - Console: matched apples, fit R2, final coefficients
  - PNG:      N_px vs cx scatter + quadratic fit (small/large groups)
  - config.yaml: updated small_coeffs, large_coeffs, small_ball_mm, large_ball_mm

Notes
-----
  - Apples are matched by ORDER: 1st committed apple = GT apple 1, etc.
  - Run with the SAME lens / height used in production (8 mm, JAI 0824-C3)
  - G1 recordings used a 6 mm lens - use those only if lens is the same
  - Multiple groups can be combined: run with --groups G1 G2 G3 ...
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
import warnings
from pathlib import Path
from typing import List, Optional, Tuple

import cv2
import numpy as np

# Silence ultralytics startup noise
logging.getLogger("ultralytics").setLevel(logging.WARNING)
warnings.filterwarnings("ignore")

# ── Repo root on sys.path so local imports work ───────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from gui.workers.tracker import AppleTracker, GradeRecord
from gui.workers.size_calibration import SizeCalibrator

log = logging.getLogger(__name__)


# =============================================================================
# GT loading
# =============================================================================

def load_gt(xlsx_path: str, groups: List[str]) -> List[float]:
    """
    Load GT caliper diameters for the requested group(s) from data/gt.xlsx.

    Columns: [Group, Apple_ID, D1, D2, Surface_Class, Average]
    Returns list of average diameters in apple order (apple 1, 2, ...).
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed. Run: pip install openpyxl")

    wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws   = wb.active
    diameters: List[float] = []
    current_group = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        grp, apple_id, d1, d2 = row[0], row[1], row[2], row[3]
        if grp:
            current_group = str(grp).strip()
        if current_group in groups and apple_id is not None and d1 is not None and d2 is not None:
            diameters.append((float(d1) + float(d2)) / 2.0)

    return diameters


# =============================================================================
# Video inference
# =============================================================================

def _prepare_input(ch1: np.ndarray, ch2, ch3, mode: str = "rb-nir1") -> np.ndarray:
    """Build the 3-channel composite the model was trained on."""
    def _gray(f):
        if f is None:
            return np.zeros(ch1.shape[:2], dtype=np.uint8)
        if f.dtype != np.uint8:
            f = cv2.normalize(f, None, 0, 255, cv2.NORM_MINMAX, cv2.CV_8U)
        return f[:, :, 0] if f.ndim == 3 else f

    if mode == "rb-nir1":
        R, B = ch1[:, :, 2], ch1[:, :, 0]
        return np.stack([R, B, _gray(ch2)], axis=2)
    elif mode == "rg-nir1":
        R, G = ch1[:, :, 2], ch1[:, :, 1]
        return np.stack([R, G, _gray(ch2)], axis=2)
    else:
        return ch1


def run_tracker_on_videos(
    paths: List[str],
    model_path: str,
    orientation: str,
    n_lanes: int,
    input_mode: str,
    conf_thr: float,
    iou_thr: float,
    device: str,
    min_dpx: float,
    skip_n: int,
) -> List[GradeRecord]:
    """
    Run YOLO+AppleTracker on a set of video files (one per channel).
    Returns the list of committed GradeRecords in sequence order.
    """
    from ultralytics import YOLO

    print(f"\nLoading model: {model_path}")
    model = YOLO(model_path)

    # Open video captures
    caps = []
    for p in paths:
        cap = cv2.VideoCapture(p)
        if not cap.isOpened():
            sys.exit(f"Cannot open video: {p}")
        caps.append(cap)
    print(f"Videos opened: {[Path(p).name for p in paths]}")

    h = int(caps[0].get(cv2.CAP_PROP_FRAME_HEIGHT))
    w = int(caps[0].get(cv2.CAP_PROP_FRAME_WIDTH))
    total_frames = int(caps[0].get(cv2.CAP_PROP_FRAME_COUNT))
    print(f"Frame size: {w}x{h}  Total frames: {total_frames}")

    # Create tracker (no calibrator - we want raw pixel sizes)
    tracker = AppleTracker(
        n_lanes    = n_lanes,
        orientation = orientation,
        exit_frac  = 0.85,
        band_half_frac = 0.025,
        entry_frac = 0.35,
        min_frames = 5,
        size_calibrator = None,   # raw pixels only
    )

    all_grades: List[GradeRecord] = []
    frame_shape = (h, w, 3)
    frame_no = 0
    t0 = time.time()

    while True:
        frames = []
        for cap in caps:
            ok, fr = cap.read()
            if not ok:
                frames.append(None)
            else:
                frames.append(fr)

        if frames[0] is None:
            break

        frame_no += 1
        if frame_no % max(1, skip_n) != 0:
            continue

        # Print progress every 50 processed frames
        if (frame_no // skip_n) % 50 == 0:
            elapsed = time.time() - t0
            pct = 100 * frame_no / max(total_frames, 1)
            print(f"  Frame {frame_no}/{total_frames}  ({pct:.0f}%)  "
                  f"Grades so far: {len(all_grades)}  "
                  f"Elapsed: {elapsed:.0f}s")

        ch1 = frames[0] if frames[0] is not None else np.zeros((h, w, 3), np.uint8)
        ch2 = frames[1] if len(frames) > 1 else None
        ch3 = frames[2] if len(frames) > 2 else None

        composite = _prepare_input(ch1, ch2, ch3, input_mode)

        results = model.track(
            source  = composite,
            tracker = "bytetrack.yaml",
            persist = True,
            conf    = conf_thr,
            iou     = iou_thr,
            device  = device,
            imgsz   = 640,
            verbose = False,
            save    = False,
        )

        _, graded = tracker.update(results[0], frame_shape)
        # Filter by minimum pixel diameter sanity check
        for rec in graded:
            if rec.size_px is not None and rec.size_px >= min_dpx:
                all_grades.append(rec)

    for cap in caps:
        cap.release()

    print(f"\nDone. {frame_no} frames processed, {len(all_grades)} apples graded.")
    return all_grades


# =============================================================================
# Quadratic fitting
# =============================================================================

def fit_group(
    cx_vals: List[float],
    dpx_vals: List[float],
    d_gt_vals: List[float],
    d_ref: float,
    label: str,
) -> Tuple[np.ndarray, float]:
    """
    Fit N_px(x) = a*x^2 + b*x + c for one group.

    Normalises each observation to d_ref mm:
        N_px_norm_i = D_px_i * (d_ref / D_mm_GT_i)

    Returns (coeffs [a,b,c], R2).
    """
    xs   = np.array(cx_vals,  dtype=float)
    ys   = np.array(dpx_vals, dtype=float)
    gts  = np.array(d_gt_vals, dtype=float)

    n_norm = ys * (d_ref / gts)

    # Quadratic fit
    coeffs  = np.polyfit(xs, n_norm, deg=2)
    y_pred  = np.polyval(coeffs, xs)
    ss_res  = np.sum((n_norm - y_pred) ** 2)
    ss_tot  = np.sum((n_norm - n_norm.mean()) ** 2)
    r2      = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0

    print(f"\n{label} group  (n={len(xs)}, d_ref={d_ref:.1f}mm)")
    print(f"  a={coeffs[0]:.6f}  b={coeffs[1]:.6f}  c={coeffs[2]:.4f}")
    print(f"  R2 = {r2:.4f}")

    return coeffs, r2


# =============================================================================
# Plot
# =============================================================================

def make_plot(
    matched: List[Tuple],
    small_coeffs, large_coeffs,
    small_ref, large_ref,
    img_width: int,
    output_path: str,
) -> None:
    """
    Scatter plot of normalised N_px vs cx, with fitted quadratic curves.
    matched: list of (cx_peak, D_px, D_mm_GT, group_label)
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("matplotlib not installed - skipping plot")
        return

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    fig.suptitle("Size Calibration — N_px(x) Quadratic Fit", fontsize=14, fontweight="bold")

    x_range = np.linspace(0, img_width, 300)

    for ax, (ref, coeffs, glabel) in zip(
        axes,
        [(small_ref, small_coeffs, "Small"), (large_ref, large_coeffs, "Large")],
    ):
        # Scatter
        pts = [(cx, dpx, dgt) for cx, dpx, dgt, gl in matched if gl == glabel]
        if pts:
            xs, ys, gts = zip(*pts)
            n_norm = np.array(ys) * (ref / np.array(gts))
            ax.scatter(xs, n_norm, alpha=0.7, s=60,
                       color="#3B82F6" if glabel == "Small" else "#F59E0B",
                       label=f"Observed  (n={len(pts)})", zorder=3)

        # Fit curve
        if coeffs is not None:
            y_fit = np.polyval(coeffs, x_range)
            ax.plot(x_range, y_fit, color="crimson", linewidth=2,
                    label=f"Quadratic fit  (d_ref={ref:.1f}mm)")

        ax.set_xlabel("Centroid column x (px)", fontsize=11)
        ax.set_ylabel(f"N_px normalised to {ref:.1f}mm", fontsize=11)
        ax.set_title(f"{glabel} apple group", fontsize=12)
        ax.legend(fontsize=10)
        ax.grid(alpha=0.3)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    print(f"\nPlot saved to: {output_path}")


# =============================================================================
# Config update
# =============================================================================

def update_config(
    config_path: str,
    small_ball_mm: float,
    small_coeffs: List[float],
    large_ball_mm: float,
    large_coeffs: List[float],
) -> None:
    """Write fitted coefficients back into config.yaml."""
    import yaml

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    se = cfg.setdefault("size_estimation", {})
    cal = se.setdefault("calibration", {})
    cal["small_ball_mm"]  = round(float(small_ball_mm), 2)
    cal["small_coeffs"]   = [round(float(v), 8) for v in small_coeffs]
    cal["large_ball_mm"]  = round(float(large_ball_mm), 2)
    cal["large_coeffs"]   = [round(float(v), 8) for v in large_coeffs]

    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

    print(f"\nConfig updated: {config_path}")
    print(f"  small_ball_mm = {small_ball_mm:.2f}  coeffs = {[round(v,6) for v in small_coeffs]}")
    print(f"  large_ball_mm = {large_ball_mm:.2f}  coeffs = {[round(v,6) for v in large_coeffs]}")


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Calibrate size estimation: fit quadratic N_px(x) from GT apple data."
    )
    ap.add_argument("--videos",  nargs="+", required=True,
                    help="Video files: ch1 [ch2 ch3] (paths for each spectral channel)")
    ap.add_argument("--gt",      required=True,
                    help="Path to gt.xlsx with caliper measurements")
    ap.add_argument("--groups",  nargs="+", default=["G1"],
                    help="Which GT session group(s) to use (e.g. G1 or G1 G2 G3)")
    ap.add_argument("--model",   required=True,
                    help="Path to YOLO model weights (.pt)")
    ap.add_argument("--config",  default="config/config.yaml",
                    help="config.yaml to update with fitted coefficients")
    ap.add_argument("--output",  default="docs/calibration.png",
                    help="Output plot path")
    ap.add_argument("--orientation", default="LR",
                    help="Conveyor orientation: LR | RL | TB | BT")
    ap.add_argument("--lanes",   type=int, default=3)
    ap.add_argument("--input-mode", default="rb-nir1",
                    help="Band combo (rb-nir1 | rg-nir1 | RGB | CH1)")
    ap.add_argument("--conf",    type=float, default=0.5)
    ap.add_argument("--iou",     type=float, default=0.45)
    ap.add_argument("--device",  default="cuda")
    ap.add_argument("--min-dpx", type=float, default=80,
                    help="Minimum bounding-box pixel size to accept (sanity filter)")
    ap.add_argument("--skip-n",  type=int, default=1,
                    help="Process every Nth frame (1=all, 2=every other, etc.)")
    ap.add_argument("--no-config-update", action="store_true",
                    help="Print coefficients but do NOT write to config.yaml")
    args = ap.parse_args()

    print("=" * 60)
    print("APPLE SIZE CALIBRATION")
    print("=" * 60)
    print(f"Groups:      {args.groups}")
    print(f"GT file:     {args.gt}")
    print(f"Model:       {args.model}")
    print(f"Videos:      {[Path(v).name for v in args.videos]}")
    print(f"Orientation: {args.orientation}   Lanes: {args.lanes}")

    # 1. Load GT
    print("\nLoading GT data...")
    gt_diameters = load_gt(args.gt, args.groups)
    if not gt_diameters:
        sys.exit(f"No GT data found for groups {args.groups}")
    print(f"  {len(gt_diameters)} apples  "
          f"range: {min(gt_diameters):.1f} - {max(gt_diameters):.1f} mm  "
          f"mean: {np.mean(gt_diameters):.1f} mm")

    # 2. Run tracker
    grades = run_tracker_on_videos(
        paths      = args.videos,
        model_path = args.model,
        orientation = args.orientation,
        n_lanes    = args.lanes,
        input_mode = args.input_mode,
        conf_thr   = args.conf,
        iou_thr    = args.iou,
        device     = args.device,
        min_dpx    = args.min_dpx,
        skip_n     = args.skip_n,
    )

    if not grades:
        sys.exit("No apples committed by tracker. Check videos, model, and tracker settings.")

    # 3. Match grades to GT by order
    n_match = min(len(grades), len(gt_diameters))
    if len(grades) != len(gt_diameters):
        print(f"\nWARNING: {len(grades)} graded vs {len(gt_diameters)} GT apples. "
              f"Using first {n_match}.")

    matched_cx   = []
    matched_dpx  = []
    matched_dgt  = []

    print(f"\n{'#':>4}  {'cx_peak':>8}  {'D_px':>7}  {'D_GT':>7}  {'r(mm/px)':>9}")
    print("-" * 45)
    for i in range(n_match):
        rec = grades[i]
        dgt = gt_diameters[i]
        if rec.size_px is None or rec.cx_peak is None:
            print(f"{i+1:>4}  {'?':>8}  {'?':>7}  {dgt:>7.1f}  (no size_px - skipped)")
            continue
        r = dgt / rec.size_px
        print(f"{i+1:>4}  {rec.cx_peak:>8.0f}  {rec.size_px:>7.1f}  {dgt:>7.1f}  {r:>9.4f}")
        matched_cx.append(rec.cx_peak)
        matched_dpx.append(rec.size_px)
        matched_dgt.append(dgt)

    if len(matched_cx) < 4:
        sys.exit(f"Only {len(matched_cx)} matched points — need at least 4 to fit a quadratic.")

    # 4. Split into small / large groups by GT median
    median_d = float(np.median(matched_dgt))
    print(f"\nMedian GT diameter: {median_d:.1f} mm")

    small_idx = [i for i, d in enumerate(matched_dgt) if d <= median_d]
    large_idx = [i for i, d in enumerate(matched_dgt) if d >  median_d]

    print(f"Small group (d <= {median_d:.1f}mm): {len(small_idx)} apples")
    print(f"Large group (d >  {median_d:.1f}mm): {len(large_idx)} apples")

    if len(small_idx) < 3 or len(large_idx) < 3:
        print("WARNING: one group has fewer than 3 points. "
              "Using single-group calibration.")
        # Fall back: use all data, treat as one group
        small_idx = list(range(len(matched_cx)))
        large_idx = list(range(len(matched_cx)))

    # Representative diameters (group mean GT)
    small_ref = float(np.mean([matched_dgt[i] for i in small_idx]))
    large_ref = float(np.mean([matched_dgt[i] for i in large_idx]))

    # 5. Fit quadratics
    print("\n" + "=" * 60)
    print("QUADRATIC FIT RESULTS")
    print("=" * 60)

    small_coeffs, small_r2 = fit_group(
        [matched_cx[i]  for i in small_idx],
        [matched_dpx[i] for i in small_idx],
        [matched_dgt[i] for i in small_idx],
        small_ref, "Small",
    )
    large_coeffs, large_r2 = fit_group(
        [matched_cx[i]  for i in large_idx],
        [matched_dpx[i] for i in large_idx],
        [matched_dgt[i] for i in large_idx],
        large_ref, "Large",
    )

    # 6. Quick self-check: compute residuals
    print("\nSelf-check residuals (calibration set):")
    cal_test = SizeCalibrator(small_ref, small_coeffs, large_ref, large_coeffs)
    errors = []
    for cx, dpx, dgt in zip(matched_cx, matched_dpx, matched_dgt):
        r = cal_test.r(cx, dpx)
        if r is not None:
            pred = dpx * r
            errors.append(pred - dgt)
    if errors:
        err = np.array(errors)
        print(f"  RMSE  = {np.sqrt(np.mean(err**2)):.2f} mm")
        print(f"  MAE   = {np.mean(np.abs(err)):.2f} mm")
        print(f"  Bias  = {np.mean(err):.2f} mm")
        all_dgt = np.array(matched_dgt)
        all_pred = all_dgt + err
        ss_res = np.sum(err**2)
        ss_tot = np.sum((all_dgt - all_dgt.mean())**2)
        r2_sys = 1 - ss_res/ss_tot
        print(f"  R2    = {r2_sys:.4f}")

    # 7. Plot
    img_w = 2048  # JAI full frame width
    matched_labelled = []
    for i in range(len(matched_cx)):
        gl = "Small" if matched_dgt[i] <= median_d else "Large"
        matched_labelled.append((matched_cx[i], matched_dpx[i], matched_dgt[i], gl))

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    make_plot(matched_labelled, small_coeffs, large_coeffs,
              small_ref, large_ref, img_w, args.output)

    # 8. Update config
    if not args.no_config_update:
        update_config(
            args.config,
            small_ball_mm = small_ref,
            small_coeffs  = small_coeffs,
            large_ball_mm = large_ref,
            large_coeffs  = large_coeffs,
        )
    else:
        print("\n--no-config-update: skipping config write.")
        print(f"  small_ball_mm={small_ref:.2f}  small_coeffs={[round(v,6) for v in small_coeffs]}")
        print(f"  large_ball_mm={large_ref:.2f}  large_coeffs={[round(v,6) for v in large_coeffs]}")

    print("\nCalibration complete.")


if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING, format="%(message)s")
    main()
