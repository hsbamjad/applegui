"""
extract_frames.py  —  Step 1: Frame Feature Extractor  (v2 — Full Redesign)
=============================================================================
Michigan State University | Apple GUI | feature/apple-size-ml branch

WHAT CHANGED (v2)
-----------------
  Old:  Custom SimpleTracker → unreliable, ghost tracks (e.g. Apple 14 in G1)
  New:  model.track() with bytetrack.yaml — same tracker used in production
        inference (gui/workers/inference_worker.py line 275).

  Old:  Raw binary masks from YOLO → jagged / sharp corners.
  New:  masks.xy polygon contour → convex hull → smooth round silhouette.

  Old:  All per-frame masks stored → ~500 MB / session pkl.
  New:  Best-quality frame mask stored as consensus + lightweight per-frame
        metadata (cx, cy, bbox, conf, quality) → ~5 MB / session pkl.

  Old:  Ghost filter = MIN_CX_TRAVEL 200 px (too aggressive for late-entrants).
  New:  Ghost filter = MIN_CX_RANGE 80 px  (cx_max − cx_min).
        Ghost tracks (stationary belt artifacts) have cx_range ≈ 10 px.
        Real apples — even late-entering ones — have cx_range ≥ 80 px.

PIPELINE (per session)
----------------------
  1.  Iterate BMP frames from Source0/<session>.
  2.  Build RB-NIR1 composite (same band combo as production model).
  3.  model.track(bytetrack.yaml, persist=True) — YOLO assigns stable IDs.
  4.  Per detection:
        a. Get polygon from masks.xy.
        b. Apply convex hull  → smooth binary mask (no sharp corners).
        c. Compute quality score  (circularity × completeness).
        d. Update per-track accumulator.
  5.  After all frames:
        a. Filter valid tracks (cx_range ≥ MIN_CX_RANGE, n_frames ≥ MIN_FRAMES,
           entered from left entry zone).
        b. Per lane: if more candidates than expected, keep top-N by n_frames.
        c. Sort kept tracks by entry frame → position in lane.
        d. Assign GT caliper measurement by entry order.
  6.  Save pickle.

USAGE (shuttle PC)
------------------
  conda run -n applegui python scripts/extract_frames.py \\
      --data_root   "D:/HA/apple_gui/images" \\
      --model_path  "D:/HA/apple_gui/models/best.pt" \\
      --gt_path     "D:/HA/apple_gui/data/gt.xlsx" \\
      --output_dir  "D:/HA/apple_gui/data/frame_features" \\
      --sessions G1 G2 G3 G4 G5 G6 G7 G8 G9 G10 G11

  Dev machine sample data:
      --data_root "D:/Haseeb/pic"

OUTPUT (pickle per session)
---------------------------
  {
    "session": "G1",
    "img_w":   2048,
    "img_h":   1536,
    "apples": [
      {
        "apple_idx":      0,          # 0-based, entry-gate order
        "lane":           0,          # 0-based lane (0=top, 1=mid, 2=bot)
        "pos_in_lane":    0,          # 0-based position within lane
        "gt_mm":          58.9,       # caliper measurement (None if unmatched)
        "track_id":       5,          # bytetrack ID
        "entry_frame":    781,        # frame_no at which apple entered entry zone
        "n_frames":       2061,       # total frames in track
        "cx_range":       1261,       # cx_max - cx_min (travel across frame)
        "best_quality":   0.84,       # quality score of consensus frame

        # Consensus mask — best-quality frame's convex-hull silhouette
        "consensus_mask": np.ndarray, # binary uint8 H×W (cropped to consensus_rect)
        "consensus_rect": [x1,y1,x2,y2], # location of consensus_mask in full frame

        # Lightweight per-frame metadata (no mask images — saves ~100× space)
        "frames": [
          {
            "frame_no":  1042,   # original BMP frame number
            "frame_idx": 0,      # 0-based sorted position
            "cx_px":     812,    # centroid x  (full-res pixels)
            "cy_px":     400,    # centroid y  (full-res pixels)
            "bbox":      [x1,y1,x2,y2],  # full-res bbox
            "conf":      0.87,   # YOLO detection confidence
            "cls_id":    1,      # 0=Fresh 1=Processing 2=Cull
            "quality":   0.72,   # circularity × completeness
            # ── Per-frame diameter estimates (pixels) — Step 2 features ──────
            "d_maxw":  float,   # M1: max projection width
            "d_sym":   float,   # M2: bilateral symmetry (Mizushima & Lu 2013)
            "d_ell":   float,   # M3: ellipse major axis
            "d_area":  float,   # M4: area sphere estimate  sqrt(4A/pi)
          },
          ...
        ],
      },
      ...
    ]
  }
"""

import argparse
import os
import re
import pickle
import sys
import numpy as np
import cv2
import openpyxl
from pathlib import Path
from ultralytics import YOLO

# Allow 'core/' package to be found when running from project root or scripts/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from core.sizing.mask_diameter import all_diameters as _all_diameters
from core.log import get_logger, configure_root

logger = get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT PATHS
#   Dev machine  : D:\\Haseeb\\pic  /  D:\\Haseeb\\Ground Truth.xlsx
#   Shuttle PC   : D:\\HA\\apple_gui\\images  /  D:\\HA\\apple_gui\\data\\gt.xlsx
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_DATA_ROOT  = r"D:\Haseeb\pic"
DEFAULT_MODEL_PATH = r"D:\HA\apple_gui\models\best.pt"
DEFAULT_GT_PATH    = r"D:\Haseeb\Ground Truth.xlsx"
DEFAULT_OUTPUT_DIR = r"D:\Haseeb\frame_features"

# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
APPLES_PER_SESSION   = 18      # 3 lanes × 6 apples
APPLES_PER_LANE      = 6

CONF_THRESHOLD       = 0.40    # YOLO detection confidence threshold
IOU_THRESHOLD        = 0.45    # YOLO NMS IoU threshold
INPUT_MODE           = "RB-nir1"  # Band composite (must match training config)
TRACKER_CFG          = "bytetrack.yaml"  # Same tracker as production inference

ENTRY_FRAC           = 0.35    # Apple valid only if first seen in left 35% of frame
MIN_FRAMES           = 50      # Track must span at least this many frames
MIN_CX_RANGE         = 80      # cx_max − cx_min must exceed this (ghost filter).
                                # Stationary belt artifacts: cx_range ≈ 10 px.
                                # Real apples (even late-entering): ≥ 80 px.
MASK_PAD             = 30      # Pixel padding around bbox for mask crop


# ─────────────────────────────────────────────────────────────────────────────
# NATURAL SORT
# ─────────────────────────────────────────────────────────────────────────────
def natural_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]



# ─────────────────────────────────────────────────────────────────────────────
# GROUND TRUTH LOADER
# ─────────────────────────────────────────────────────────────────────────────
def load_gt(gt_path: str, session: str) -> list:
    """
    Load caliper GT measurements for one session from the Excel file.

    File format (confirmed):
      Sheet: 'Sheet1'  (single sheet, all sessions)
      Columns: Session | Apple# | D1 | D2 | SurfaceClass | Average
      'Session' column only has a value (e.g. 'G1') in the FIRST row of each
      session block; remaining rows of that session have None in that column.
      'Average' = (D1 + D2) / 2  in mm  — this is the GT diameter.

    Returns a list of Average floats in the order they appear for the session.
    Falls back gracefully (empty list + warning) on any error.
    """
    try:
        wb = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)

        # Find the sheet — prefer a sheet named after the session, fall back to Sheet1
        target_sheet = None
        for name in wb.sheetnames:
            if name.strip().upper() == session.strip().upper():
                target_sheet = name
                break
        if target_sheet is None and "Sheet1" in wb.sheetnames:
            target_sheet = "Sheet1"
        if target_sheet is None:
            logger.warning("[GT] No usable sheet found in %s (sheets: %s)",
                           gt_path, wb.sheetnames)
            wb.close()
            return []

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row
        wb.close()

        # ── Case 1: sheet is named after the session (old format) ─────────────
        # All rows belong to this session; grab every numeric cell in order.
        if target_sheet.strip().upper() == session.strip().upper():
            vals = []
            for row in rows:
                for cell in row:
                    if cell is not None:
                        try:
                            vals.append(float(cell))
                        except (TypeError, ValueError):
                            pass
            logger.info("[GT] Loaded %d values from sheet '%s'", len(vals), target_sheet)
            return vals

        # ── Case 2: Sheet1 multi-session format ───────────────────────────────
        # Col 0 = session label (G1/G2/…, None for continuation rows)
        # Col 1 = apple number (1-18)
        # Col 5 = Average diameter (mm)  ← the GT value we want
        AVERAGE_COL = 5
        in_session = False
        vals = []
        for row in rows:
            label = row[0]  # session column
            if label is not None:
                # New session starting — check if it matches our target
                in_session = str(label).strip().upper() == session.strip().upper()
            if not in_session:
                continue
            # Extract Average column
            if len(row) > AVERAGE_COL and row[AVERAGE_COL] is not None:
                try:
                    vals.append(float(row[AVERAGE_COL]))
                except (TypeError, ValueError):
                    pass

        logger.info("[GT] Loaded %d values for session '%s' from '%s'  (D1+D2)/2 averages in mm",
                    len(vals), session, target_sheet)
        return vals

    except Exception as exc:
        logger.warning("[GT] Failed to load %s: %s", gt_path, exc)
        return []




# ─────────────────────────────────────────────────────────────────────────────
# CHANNEL COMPOSITE  (matches training config INPUT_MODE = "RB-nir1")
# ─────────────────────────────────────────────────────────────────────────────
def build_model_input(src0_bgr: np.ndarray, src1_gray: np.ndarray) -> np.ndarray:
    """
    Builds the 3-channel uint8 image fed to YOLO.
    Matches the band combo used in RealInferenceWorker._prepare_input('rb-nir1').

      Channel 0 : Red   (src0[:,:,2])
      Channel 1 : Blue  (src0[:,:,0])
      Channel 2 : NIR1  (src1 grayscale)

    Handles all Source1 image formats robustly:
      - (H, W)    — already 2-D grayscale (most common)
      - (H, W, 1) — 1-channel BMP (some cameras/OS save this way)
      - (H, W, 3) — 3-channel BGR accidentally loaded without GRAYSCALE flag
    """
    R = src0_bgr[:, :, 2]
    B = src0_bgr[:, :, 0]

    if src1_gray.ndim == 2:
        N1 = src1_gray                                          # already 2-D
    elif src1_gray.ndim == 3 and src1_gray.shape[2] == 1:
        N1 = src1_gray[:, :, 0]                                # squeeze single channel
    elif src1_gray.ndim == 3 and src1_gray.shape[2] >= 3:
        N1 = cv2.cvtColor(src1_gray, cv2.COLOR_BGR2GRAY)       # convert BGR → gray
    else:
        N1 = src1_gray.reshape(src1_gray.shape[:2])             # fallback: force 2-D

    return np.stack([R, B, N1], axis=2)



# ─────────────────────────────────────────────────────────────────────────────
# MASK PROCESSING  (ellipse-fit based — smooth, round, physically correct)
# ─────────────────────────────────────────────────────────────────────────────
def make_smooth_mask(
    poly:    np.ndarray,    # YOLO masks.xy contour  (N x 2 float, image coords)
    x1: int, y1: int,
    x2: int, y2: int,
    img_w: int, img_h: int,
) -> tuple:
    """
    Fit an ellipse to the YOLO polygon and return a smooth ellipse mask.

    Why ellipse?
      - Apples are spherical -> 2-D silhouette IS an ellipse.
      - cv2.fitEllipse averages all polygon vertices -> robust to YOLO noise.
      - Ellipse axes (axis_a, axis_b) are directly the two diameters D1, D2.
      - No straight edges: perfectly smooth boundary (unlike convexHull polygon).

    Steps:
      1. Compute convex hull of YOLO polygon (removes any concavities/jaggies).
      2. Fit an ellipse to the hull points using cv2.fitEllipse().
      3. Draw the filled ellipse as the binary mask.
      4. Quality = circularity of the resulting ellipse mask x completeness.

    Returns:
      smooth_mask  : uint8 binary crop (crop_h x crop_w), 255 = apple
      crop_rect    : [mx1, my1, mx2, my2] in full-frame pixel coords
      quality      : float in [0, 1]
      ellipse_abs  : (center_x, center_y, axis_a, axis_b, angle) in full-frame
                     coords — or None if fitting failed
    """
    if poly is None or len(poly) < 5:          # fitEllipse needs >= 5 points
        return None, None, 0.0, None

    # ── Crop region: bbox + padding ──────────────────────────────────────────
    mx1 = max(0,     x1 - MASK_PAD)
    my1 = max(0,     y1 - MASK_PAD)
    mx2 = min(img_w, x2 + MASK_PAD)
    my2 = min(img_h, y2 + MASK_PAD)
    crop_w = mx2 - mx1
    crop_h = my2 - my1

    # ── Convex hull in full-frame coords (removes concavities / jaggies) ─────
    poly_int = poly.astype(np.int32)
    hull     = cv2.convexHull(poly_int)         # still in full-frame coords

    # ── Fit ellipse in full-frame coords ─────────────────────────────────────
    try:
        ellipse_full = cv2.fitEllipse(hull)     # ((cx,cy), (ma,mi), angle)
    except cv2.error:
        # Degenerate polygon — fall back to bbox circle
        cx_f = (x1 + x2) / 2.0
        cy_f = (y1 + y2) / 2.0
        r    = max((x2 - x1), (y2 - y1)) / 2.0
        ellipse_full = ((cx_f, cy_f), (r * 2, r * 2), 0.0)

    (cx_f, cy_f), (axis_ma, axis_mi), angle = ellipse_full
    # axis_ma = major axis length (full diameter), axis_mi = minor axis length

    # ── Draw filled ellipse in crop coordinates ───────────────────────────────
    cx_crop = cx_f - mx1
    cy_crop = cy_f - my1
    ell_crop = ((cx_crop, cy_crop), (axis_ma, axis_mi), angle)

    smooth_mask = np.zeros((crop_h, crop_w), dtype=np.uint8)
    try:
        cv2.ellipse(smooth_mask, ell_crop, 255, -1)   # filled ellipse
    except cv2.error:
        return None, None, 0.0, None

    if not np.any(smooth_mask):
        return None, None, 0.0, None

    # ── Quality = axis_ratio × completeness ──────────────────────────────────
    # axis ratio: 1.0 = perfect circle, <1 = elongated
    circ = float(min(axis_ma, axis_mi) / (max(axis_ma, axis_mi) + 1e-6))

    # ── Completeness: apple fully in frame = center is at least half-width from edge ──
    # Old check: x1<=2 fails for Apple 14 which enters very gradually
    # (its bbox x1 stays > 2 even when only a sliver is visible).
    # New check: center must be at least one apple-radius from every edge.
    half_w = (x2 - x1) / 2.0
    half_h = (y2 - y1) / 2.0
    margin = 10   # small pixel buffer
    is_complete = (
        cx_f - half_w > margin and
        cx_f + half_w < img_w - margin and
        cy_f - half_h > margin and
        cy_f + half_h < img_h - margin
    )
    completeness = 1.0 if is_complete else 0.5

    quality = circ * completeness


    # ── Return full-frame ellipse params ─────────────────────────────────────
    ellipse_abs = (cx_f, cy_f, axis_ma, axis_mi, angle)

    return smooth_mask, [mx1, my1, mx2, my2], quality, ellipse_abs


def compute_consensus_ellipse(
    frame_ellipses: list,   # list of (cx, cy, axis_ma, axis_mi, angle) per frame
    frame_cx_vals:  list,   # parallel list of per-frame cx_px values
    cx_min: float,          # track cx_min (left-most centroid)
    cx_max: float,          # track cx_max (right-most centroid)
    img_w: int,
    img_h: int,
    central_frac: float = 0.20,  # exclude first+last 20% of traversal each side
) -> tuple:
    """
    Compute a robust consensus ellipse using only frames where the apple is
    fully in view — defined as frames where cx is within the CENTRAL portion
    of the apple's traversal.

    Why cx-based rather than bbox-edge-based?
    -----------------------------------------
    Apple 14 enters gradually: its bbox x1 is never 0 (YOLO clips to the
    visible portion), so the old 'touches edge' check failed to exclude its
    957 entry frames where only a slice was visible, corrupting the median.

    With cx-range filtering:
      cx_low  = cx_min + 0.20 * cx_range  →  skip first 20% of traversal
      cx_high = cx_max - 0.20 * cx_range  →  skip last 20% of traversal
    This robustly excludes partial-entry and partial-exit frames for ALL apples.

    Returns:
      consensus_params dict, rendered binary mask, crop_rect [mx1,my1,mx2,my2]
    """
    if not frame_ellipses:
        return None, None, None

    cx_range = cx_max - cx_min
    if cx_range > 0 and len(frame_cx_vals) == len(frame_ellipses):
        cx_low  = cx_min + central_frac * cx_range
        cx_high = cx_max - central_frac * cx_range
        central_ells = [
            e for e, cxv in zip(frame_ellipses, frame_cx_vals)
            if cx_low <= cxv <= cx_high
        ]
    else:
        central_ells = frame_ellipses

    # Fall back to all frames if too few central frames
    if len(central_ells) < 10:
        central_ells = frame_ellipses

    arr = np.array(central_ells, dtype=np.float32)  # (N, 5)
    cx_med    = float(np.median(arr[:, 0]))
    cy_med    = float(np.median(arr[:, 1]))
    ma_med    = float(np.median(arr[:, 2]))
    mi_med    = float(np.median(arr[:, 3]))
    # Angle: circular median (angles wrap at 180 deg)
    angles_rad = np.deg2rad(arr[:, 4])
    ang_med    = float(np.rad2deg(np.arctan2(
        np.median(np.sin(2 * angles_rad)),
        np.median(np.cos(2 * angles_rad))
    ) / 2.0))

    # ── Render the consensus ellipse into a padded crop ───────────────────────
    pad   = MASK_PAD
    mx1   = max(0,     int(cx_med - ma_med / 2) - pad)
    my1   = max(0,     int(cy_med - ma_med / 2) - pad)
    mx2   = min(img_w, int(cx_med + ma_med / 2) + pad)
    my2   = min(img_h, int(cy_med + ma_med / 2) + pad)
    crop_w = mx2 - mx1
    crop_h = my2 - my1

    if crop_w <= 0 or crop_h <= 0:
        return None, None, None

    cx_crop = cx_med - mx1
    cy_crop = cy_med - my1

    mask = np.zeros((crop_h, crop_w), dtype=np.uint8)
    try:
        cv2.ellipse(mask, ((cx_crop, cy_crop), (ma_med, mi_med), ang_med), 255, -1)
    except cv2.error:
        return None, None, None

    consensus_params = {
        "cx":     cx_med,
        "cy":     cy_med,
        "axis_a": ma_med,          # major axis (pixels, full diameter)
        "axis_b": mi_med,          # minor axis (pixels, full diameter)
        "angle":  ang_med,
        "n_frames_used": len(central_ells),
    }
    return consensus_params, mask, [mx1, my1, mx2, my2]



# ─────────────────────────────────────────────────────────────────────────────
# LANE ASSIGNMENT
# ─────────────────────────────────────────────────────────────────────────────
def assign_lanes(valid_tracks: list, img_h: int, n_lanes: int = 3) -> list:
    """
    Assign each valid track to a lane (0-based) based on the median cy
    across its frames.

    Uses simple equal-height binning of the frame.
    """
    lane_h = img_h / n_lanes
    for td in valid_tracks:
        cy_vals = [fr["cy_px"] for fr in td["frames"]]
        cy_med  = float(np.median(cy_vals))
        lane    = int(cy_med / lane_h)
        lane    = max(0, min(n_lanes - 1, lane))
        td["lane"] = lane
    return valid_tracks


def select_and_order_tracks(
    valid_tracks: list,
    n_lanes:      int = 3,
    apples_per_lane: int = 6,
) -> list:
    """
    For each lane:
      1. If more candidates than expected (apples_per_lane), keep the top-N
         by frame count (longest-tracked = most likely real apples).
      2. Sort kept tracks by entry_frame ascending -> pos_in_lane within lane.

    apple_idx uses the FIXED POSITIONAL CONVENTION confirmed by the user:

        apple_idx = pos_in_lane * n_lanes + lane

    This means:
        Apple 0 (displayed as #1) -> Lane 0, Pos 0
        Apple 1 (displayed as #2) -> Lane 1, Pos 0
        Apple 2 (displayed as #3) -> Lane 2, Pos 0
        Apple 3 (displayed as #4) -> Lane 0, Pos 1
        Apple 4 (displayed as #5) -> Lane 1, Pos 1
        Apple 5 (displayed as #6) -> Lane 2, Pos 1
        ...

    This is INDEPENDENT of entry_frame order. Entry frame is only used to
    determine which tracked apple occupies which belt position (pos_in_lane).
    Even if lane 2 enters the camera before lane 1, lane 1 is always #5 and
    lane 2 is always #6. If an apple falls off the belt mid-traversal, its
    slot stays fixed and doesn't shift the numbering of others.
    """
    assigned = []
    for lane_idx in range(n_lanes):
        lane_tracks = [td for td in valid_tracks if td["lane"] == lane_idx]

        # Trim excess candidates
        if len(lane_tracks) > apples_per_lane:
            lane_tracks.sort(key=lambda t: t["n_frames"], reverse=True)
            discarded = lane_tracks[apples_per_lane:]
            lane_tracks = lane_tracks[:apples_per_lane]
            logger.warning("Lane %d: %d candidates -> kept top %d by frame count (discarded %d)",
                           lane_idx, len(lane_tracks)+len(discarded),
                           apples_per_lane, len(discarded))
        elif len(lane_tracks) < apples_per_lane:
            logger.warning("Lane %d: only %d tracks (expected %d) -- missing apples?",
                           lane_idx, len(lane_tracks), apples_per_lane)

        # Sort by entry frame within lane -> positional order (first to enter = pos 0)
        lane_tracks.sort(key=lambda t: t["entry_frame"])
        for pos, td in enumerate(lane_tracks):
            td["pos_in_lane"] = pos
        assigned.extend(lane_tracks)

    # apple_idx = POSITIONAL CONVENTION  (pos_in_lane * n_lanes + lane)
    # This matches the GT Excel numbering and the physical labels on the video.
    for td in assigned:
        td["apple_idx"] = td["pos_in_lane"] * n_lanes + td["lane"]

    # Sort for display: by apple_idx ascending
    assigned.sort(key=lambda t: t["apple_idx"])
    return assigned


def assign_gt(assigned_tracks: list, gt_list: list) -> None:
    """
    Assign GT caliper measurements using the FIXED POSITIONAL CONVENTION.

    GT Excel ordering (confirmed by user):
        Apple 1  = Lane 0, Pos 0   (gt_list[0])
        Apple 2  = Lane 1, Pos 0   (gt_list[1])
        Apple 3  = Lane 2, Pos 0   (gt_list[2])
        Apple 4  = Lane 0, Pos 1   (gt_list[3])
        ...
        Apple N  = Lane (N-1)%3, Pos (N-1)//3

    apple_idx = pos_in_lane * n_lanes + lane  (set by select_and_order_tracks)
    This directly indexes into gt_list.
    """
    for td in assigned_tracks:
        idx = td["apple_idx"]          # pos_in_lane*3 + lane
        td["gt_mm"] = gt_list[idx] if idx < len(gt_list) else None


# ─────────────────────────────────────────────────────────────────────────────
# SESSION PROCESSOR
# ─────────────────────────────────────────────────────────────────────────────
def process_session(
    session:     str,
    data_root:   str,
    model:       YOLO,
    gt_path:     str,
    output_dir:  str,
    device:      str = "cuda",
) -> None:
    """Run model.track() on all frames of one session and save a pickle."""

    src0_dir = Path(data_root) / "Source0" / session
    src1_dir = Path(data_root) / "Source1" / session

    if not src0_dir.exists():
        logger.warning("[FAIL] Source0 folder not found: %s", src0_dir)
        return

    src0_files = sorted(
        [f for f in src0_dir.iterdir() if f.suffix.lower() == ".bmp"],
        key=lambda f: natural_key(f.name),
    )
    if not src0_files:
        logger.warning("[FAIL] No BMP files in %s", src0_dir)
        return

    logger.info("-" * 60)
    logger.info("Session %s  |  %d frames", session, len(src0_files))

    # Detect image dimensions from first frame
    probe = cv2.imread(str(src0_files[0]))
    if probe is None:
        logger.warning("[FAIL] Cannot read first frame: %s", src0_files[0])
        return
    img_h, img_w = probe.shape[:2]
    logger.info("Image size: %dx%d", img_w, img_h)
    del probe

    entry_px = int(img_w * ENTRY_FRAC)   # x threshold for valid entry

    # ── Per-track accumulator ─────────────────────────────────────────────────
    # track_data[bytetrack_id] = dict with running state
    track_data: dict[int, dict] = {}

    # ── Main frame loop ───────────────────────────────────────────────────────
    for frame_idx, src0_path in enumerate(src0_files):
        frame_no = int(src0_path.stem.split("_")[-1])

        src0 = cv2.imread(str(src0_path))
        if src0 is None:
            continue

        # Source1 (NIR1) — optional; fall back to zeros if missing
        src1_path = src1_dir / src0_path.name
        if src1_path.exists():
            src1 = cv2.imread(str(src1_path), cv2.IMREAD_GRAYSCALE)
            if src1 is None:
                src1 = np.zeros((img_h, img_w), dtype=np.uint8)
        else:
            src1 = np.zeros((img_h, img_w), dtype=np.uint8)

        frame_input = build_model_input(src0, src1)

        # ── YOLO tracking ─────────────────────────────────────────────────────
        results = model.track(
            source       = frame_input,
            tracker      = TRACKER_CFG,
            persist      = True,
            conf         = CONF_THRESHOLD,
            iou          = IOU_THRESHOLD,
            imgsz        = 640,
            device       = device,
            retina_masks = True,
            verbose      = False,
            save         = False,
        )

        res = results[0]
        if res.boxes is None or res.boxes.id is None:
            continue

        boxes_np   = res.boxes.cpu().numpy()
        masks_obj  = res.masks          # may be None if model has no seg head
        track_ids  = boxes_np.id.astype(int)
        cls_ids    = boxes_np.cls.astype(int)
        confs      = boxes_np.conf
        xyxys      = boxes_np.xyxy

        for i in range(len(track_ids)):
            tid  = int(track_ids[i])
            cls  = int(cls_ids[i])
            conf = float(confs[i])
            x1, y1, x2, y2 = map(int, xyxys[i])
            cx = (x1 + x2) // 2
            cy = (y1 + y2) // 2

            # ── Smooth mask ───────────────────────────────────────────────────
            if masks_obj is not None and i < len(masks_obj.xy):
                poly = masks_obj.xy[i]
            else:
                poly = None

            smooth_mask, crop_rect, quality, ellipse_abs = make_smooth_mask(
                poly, x1, y1, x2, y2, img_w, img_h
            )

            # ── Update accumulator ────────────────────────────────────────────
            if tid not in track_data:
                track_data[tid] = {
                    "frames":         [],
                    "ellipses":       [],   # per-frame ellipse params (full-frame)
                    "completeness":   [],   # per-frame: True = apple fully in frame
                    "cx_min":         cx,
                    "cx_max":         cx,
                    "entry_cx":       cx,
                    "entry_frame":    frame_no,
                    "best_quality":   -1.0,
                    "best_mask":      None,
                    "best_rect":      None,
                }
            else:
                td = track_data[tid]
                td["cx_min"] = min(td["cx_min"], cx)
                td["cx_max"] = max(td["cx_max"], cx)

            td = track_data[tid]

            # Collect per-frame ellipse for consensus
            if ellipse_abs is not None:
                td["ellipses"].append(ellipse_abs)
                # Same center-radius completeness logic as make_smooth_mask
                half_w = (x2 - x1) / 2.0
                half_h = (y2 - y1) / 2.0
                margin = 10
                complete = (
                    cx - half_w > margin and
                    cx + half_w < img_w - margin and
                    cy - half_h > margin and
                    cy + half_h < img_h - margin
                )
                td["completeness"].append(complete)

            # Keep best single-frame mask as backup
            if quality > td["best_quality"] and smooth_mask is not None:
                td["best_quality"] = quality
                td["best_mask"]    = smooth_mask.copy()
                td["best_rect"]    = crop_rect

            # ── All 4 diameter methods from raw YOLO polygon ─────────────────
            # Computed on the CONVEX HULL of the raw polygon (same hull used
            # in make_smooth_mask). Using the raw poly (not the rendered ellipse
            # mask) gives independent estimates — essential for Step 2 ML.
            if poly is not None and len(poly) >= 5:
                # Render hull to a small binary mask for mask_diameter functions
                hull_pts = cv2.convexHull(
                    poly.astype(np.int32).reshape(-1, 1, 2)
                )  # shape (N,1,2)
                # Bounding box of hull for crop
                hx, hy, hw, hh = cv2.boundingRect(hull_pts)
                hx1 = max(0, hx - 4); hy1 = max(0, hy - 4)
                hx2 = min(img_w, hx + hw + 4); hy2 = min(img_h, hy + hh + 4)
                cw, ch = hx2 - hx1, hy2 - hy1
                if cw > 0 and ch > 0:
                    hull_crop = np.zeros((ch, cw), dtype=np.uint8)
                    hull_shifted = hull_pts.copy()
                    hull_shifted[:, :, 0] -= hx1
                    hull_shifted[:, :, 1] -= hy1
                    cv2.fillPoly(hull_crop, [hull_shifted.reshape(-1, 2)], 255)
                    diam = _all_diameters(hull_crop, angle_step=5)
                    d_maxw = diam["d_maxwidth"]
                    d_sym  = diam["d_symmetry"]
                    d_ell  = diam["d_ellipse"]
                    d_area = diam["d_area"]
                else:
                    d_maxw = d_sym = d_ell = d_area = 0.0
            else:
                d_maxw = d_sym = d_ell = d_area = 0.0

            # Lightweight per-frame record
            td["frames"].append({
                "frame_no":  frame_no,
                "frame_idx": frame_idx,
                "cx_px":     cx,
                "cy_px":     cy,
                "bbox":      [x1, y1, x2, y2],
                "conf":      conf,
                "cls_id":    cls,
                "quality":   quality,
                # Per-frame ellipse axes (pixels) — from consensus ellipse
                "ell_a":     ellipse_abs[2] if ellipse_abs else None,
                "ell_b":     ellipse_abs[3] if ellipse_abs else None,
                "ell_angle": ellipse_abs[4] if ellipse_abs else None,
                # Per-frame diameter estimates (pixels) — M1-M4, Step 2 features
                "d_maxw":    d_maxw,
                "d_sym":     d_sym,
                "d_ell":     d_ell,
                "d_area":    d_area,
            })

        if frame_idx % 200 == 0 and frame_idx > 0:
            logger.debug("... frame %d/%d  active tracks: %d",
                         frame_idx, len(src0_files), len(track_data))

    logger.info("Raw bytetrack IDs seen: %d", len(track_data))

    # ── Filter valid tracks ───────────────────────────────────────────────────
    valid = []
    for tid, td in track_data.items():
        n_frames  = len(td["frames"])
        cx_range  = td["cx_max"] - td["cx_min"]
        entry_cx  = td["entry_cx"]

        if n_frames  < MIN_FRAMES:
            continue   # too few frames — phantom detection
        if cx_range  < MIN_CX_RANGE:
            continue   # barely moved — stationary belt artifact (ghost)
        if entry_cx  > entry_px:
            continue   # first seen after entry zone — not a full traversal

        td["n_frames"]  = n_frames
        td["cx_range"]  = cx_range
        td["track_id"]  = tid
        valid.append(td)

    logger.info("Valid tracks after filtering: %d  (expected %d)", len(valid), APPLES_PER_SESSION)

    # ── Lane assignment ───────────────────────────────────────────────────────
    assign_lanes(valid, img_h, n_lanes=3)

    # ── Select & order tracks ─────────────────────────────────────────────────
    assigned = select_and_order_tracks(valid, n_lanes=3, apples_per_lane=APPLES_PER_LANE)

    # ── GT assignment ─────────────────────────────────────────────────────────
    gt_list = load_gt(gt_path, session) if gt_path else []
    assign_gt(assigned, gt_list)

    # ── Build output records (compute consensus ellipse per apple) ─────────────
    committed_apples = []
    for td in assigned:
        gt_mm  = td.get("gt_mm")
        gt_str = f"{gt_mm:.1f}mm" if gt_mm is not None else "None"

        # ── Consensus ellipse: median over CENTRAL traversal frames ──────────
        # Use cx values from per-frame records to filter to the central 60%
        # of the apple's traversal. This robustly excludes entry/exit frames
        # where only a partial apple is visible (fixes Apple 14 issue).
        frame_cx_list = [fr["cx_px"] for fr in td["frames"]
                         if fr.get("ell_a") is not None]  # only frames with ellipse
        consensus_params, consensus_mask, consensus_rect = compute_consensus_ellipse(
            td["ellipses"],
            frame_cx_list,
            td["cx_min"],
            td["cx_max"],
            img_w,
            img_h,
        )

        # Fall back to best single-frame mask if consensus failed
        if consensus_mask is None:
            consensus_mask = td["best_mask"]
            consensus_rect = td["best_rect"]

        # Quality: median of CENTRAL-traversal frame quality values only
        # (same 20-80% cx filter — excludes partial entry/exit frames)
        cx_range_td = td["cx_max"] - td["cx_min"]
        cx_low_q  = td["cx_min"] + 0.20 * cx_range_td
        cx_high_q = td["cx_max"] - 0.20 * cx_range_td
        central_qualities = [
            fr["quality"] for fr in td["frames"]
            if cx_low_q <= fr["cx_px"] <= cx_high_q and fr["quality"] > 0
        ]
        if central_qualities:
            best_q = float(np.median(central_qualities))
        else:
            best_q = float(np.median([f["quality"] for f in td["frames"] if f["quality"] > 0])) \
                     if any(f["quality"] > 0 for f in td["frames"]) else td["best_quality"]


        ell_a = consensus_params["axis_a"] if consensus_params else None
        ell_b = consensus_params["axis_b"] if consensus_params else None
        ell_info = (f"ell={ell_a:.1f}x{ell_b:.1f}px" if ell_a else "no_ell")

        logger.info("Apple %2d  lane=%d pos=%d  frames=%4d  cx_range=%4dpx  Q=%.3f  %s  gt=%s",
                    td['apple_idx']+1, td['lane'], td['pos_in_lane'],
                    td['n_frames'], td['cx_range'], best_q, ell_info, gt_str)

        committed_apples.append({
            "apple_idx":        td["apple_idx"],
            "lane":             td["lane"],
            "pos_in_lane":      td["pos_in_lane"],
            "gt_mm":            td.get("gt_mm"),
            "track_id":         td["track_id"],
            "entry_frame":      td["entry_frame"],
            "n_frames":         td["n_frames"],
            "cx_range":         td["cx_range"],
            "best_quality":     best_q,
            # Consensus ellipse — median over all frames
            "consensus_params": consensus_params,
            "consensus_mask":   consensus_mask,
            "consensus_rect":   consensus_rect,
            # Per-frame lightweight metadata (includes ell_a, ell_b, ell_angle)
            "frames":           td["frames"],
        })

    logger.info("Total assigned: %d  (expected %d)", len(committed_apples), APPLES_PER_SESSION)

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)
    out_path = Path(output_dir) / f"{session}.pkl"
    payload = {
        "session": session,
        "img_w":   img_w,
        "img_h":   img_h,
        "apples":  committed_apples,
    }
    with open(out_path, "wb") as f:
        pickle.dump(payload, f)
    logger.info("Saved -> %s", out_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Step 1: Extract per-apple features from G1-G11 sessions (v2)"
    )
    parser.add_argument("--data_root",  default=DEFAULT_DATA_ROOT,
                        help="Root containing Source0/Source1 folders")
    parser.add_argument("--model_path", default=DEFAULT_MODEL_PATH,
                        help="Path to YOLO best.pt")
    parser.add_argument("--gt_path",    default=DEFAULT_GT_PATH,
                        help="Path to GT Excel file (Ground Truth.xlsx)")
    parser.add_argument("--output_dir", default=DEFAULT_OUTPUT_DIR,
                        help="Directory to save .pkl files")
    parser.add_argument("--sessions",   nargs="+",
                        default=[f"G{i}" for i in range(1, 12)],
                        help="Sessions to process, e.g. G1 G2 G10")
    parser.add_argument("--device",     default="cuda",
                        help="Inference device: cuda or cpu")
    args = parser.parse_args()

    configure_root()
    logger.info("Loading model: %s", args.model_path)
    model = YOLO(args.model_path)
    model.to(args.device)
    logger.info("Model loaded. Device: %s", args.device)
    logger.info("Tracker: %s  |  Input mode: %s", TRACKER_CFG, INPUT_MODE)
    logger.info("Sessions to process: %s", args.sessions)

    for session in args.sessions:
        process_session(
            session    = session,
            data_root  = args.data_root,
            model      = model,
            gt_path    = args.gt_path,
            output_dir = args.output_dir,
            device     = args.device,
        )

    logger.info("All sessions complete.")


if __name__ == "__main__":
    main()
